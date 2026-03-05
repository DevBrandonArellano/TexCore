import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def format_excel_worksheet(ws):
    """Aplica formato visual estilizado a una hoja de Excel."""
    # Estilos Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # Aplicar al encabezado (asumiendo que está en la fila 1)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Autoajuste de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 50) # Cap a 50

def dataframe_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Reporte") -> bytes:
    """Convierte un DataFrame de pandas en bytes de Excel (.xlsx) estilizado."""
    output = BytesIO()
    
    # Escribir con pandas
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]
        format_excel_worksheet(worksheet)
        
    return output.getvalue()
